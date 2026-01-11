import os
import unicodedata
from typing import Dict, cast, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Ensure openpyxl Objects are strictly typed if necessary
# Pylance strict mode might need help with openpyxl's dynamic attributes


def ensure_dir(directory: str) -> None:
    """Ensures that a directory exists, creating it if necessary."""
    if directory and not os.path.exists(directory):
        try:
            os.makedirs(directory)
            print(f"Created directory: {directory}")
        except Exception as e:
            print(f"Error creating directory {directory}: {e}")


def load_uid_map(file_path: str) -> Dict[str, str]:
    """Loads UID to Name mapping from an Excel file."""
    uid_map: Dict[str, str] = {}
    if not os.path.exists(file_path):
        print(f"Warning: UID map file not found at {file_path}")
        return uid_map

    try:
        # Pylance Strict: read_excel returns DataFrame | TextFileReader. Cast to DataFrame.
        # type: ignore to suppress "partially unknown" warnings from pandas stubs
        uid_df = cast(pd.DataFrame, pd.read_excel(file_path))  # type: ignore
        # Normalize column names
        uid_df.columns = [str(c).lower() for c in uid_df.columns]
        if "uid" in uid_df.columns and "name" in uid_df.columns:
            # Ensure strings for matching
            uid_df["uid"] = uid_df["uid"].astype(str)
            uid_map = dict(zip(uid_df["uid"], uid_df["name"]))
        else:
            print(f"Warning: Columns 'uid' and 'name' not found in {file_path}")
    except Exception as e:
        print(f"Error reading UID map from {file_path}: {e}")

    return uid_map


def _get_char_width(char: str) -> float:
    """
    计算单个字符的视觉宽度（估算值）。
    宽字符（F/W/A）计为 1.8，其他计为 1.1。
    """
    if unicodedata.east_asian_width(char) in ("F", "W", "A"):
        return 1.8
    return 1.1


def apply_excel_styles(filename: str) -> None:
    """
    应用自定义 Excel 样式：
    - 字体：Microsoft YaHei, 12pt
    - 对齐：居中
    - 边框：内容单元格所有边为细线
    - 标题：加粗
    - 固定行高（18）
    - 自适应列宽
    """
    try:
        wb = load_workbook(filename)

        # 预定义样式对象，避免重复创建
        font_normal = Font(name="Microsoft YaHei", size=12)
        font_bold = Font(name="Microsoft YaHei", size=12, bold=True)
        alignment_style = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for ws in wb.worksheets:
            # 记录每列的最大宽度 {col_idx (1-based): max_width}
            col_widths: Dict[int, float] = {}

            for row_raw in ws.iter_rows():
                # Cast row to Tuple[Cell, ...] to help type checking
                row = cast(Tuple[Cell, ...], row_raw)

                if not row:
                    continue

                # 设置行高
                # Ensure row index is int
                ws.row_dimensions[row[0].row].height = 18

                for cell in row:
                    # 应用样式
                    cell.alignment = alignment_style
                    cell.border = thin_border
                    cell.font = font_bold if cell.row == 1 else font_normal

                    # 计算内容宽度
                    cell_value = cell.value
                    if cell_value is not None:
                        cell_text = str(cell_value)
                        # 使用 sum() 和 生成器表达式简化逻辑
                        cell_len = sum(_get_char_width(char) for char in cell_text)

                        current_max = col_widths.get(cell.column, 0.0)
                        col_widths[cell.column] = max(current_max, cell_len)

            # 应用列宽
            for col_idx, length in col_widths.items():
                col_letter = get_column_letter(col_idx)

                # 默认最小宽度
                final_width = max(length + 4, 10)

                # 特殊处理：获取该列标题
                # 注意：假设第一行为标题行
                header_cell = ws.cell(row=1, column=col_idx)
                header_val = header_cell.value

                # 针对 "题目名称" 列的特殊调整
                if header_val and str(header_val).strip() == "题目名称":
                    final_width = min(length + 8, 80)

                ws.column_dimensions[col_letter].width = final_width

        wb.save(filename)
        print("Styles applied successfully.")

    except PermissionError:
        print(
            f"Error: Permission denied saving '{filename}'. Please close the file if it is open."
        )
    except Exception as e:
        print(f"Error applying styles: {e}")
